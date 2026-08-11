"""Fail-closed parser for the standard WeChat Pay XLSX statement export."""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timezone, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from hashlib import sha256
from pathlib import Path
import re

FORMAT_ID = "wechat-pay-xlsx"
TIMEZONE = "Asia/Shanghai"
_TZ = timezone(timedelta(hours=8))
_ZERO = Decimal("0.00")
_MONEY = re.compile(r"[-+]?\s*[¥￥]?\s*[\d,]+(?:\.\d+)?")
_ORDER_NOISE = re.compile(r"(?:订单号|商户单号|交易单号|流水号|单号)\s*[:：#]?\s*[A-Za-z0-9_-]{6,}", re.I)
_NOTE_NOISE = re.compile(r"(?:订单支付|二维码收款)", re.I)
_CARD_NOISE = re.compile(r"(?<!\d)(?:\d[ -]?){11,19}(?!\d)")
_EMAIL_NOISE = re.compile(r"\b[^\s@]+@[^\s@]+\.[^\s@]+\b", re.I)

_HEADERS = {
    "occurred_at": ("交易时间",),
    "transaction_type": ("交易类型",),
    "counterparty": ("交易对方",),
    "product": ("商品", "商品名称"),
    "direction": ("收/支", "收支"),
    "amount": ("金额(元)", "金额（元）", "金额"),
    "payment_method": ("支付方式",),
    "status": ("当前状态", "交易状态"),
    "transaction_id": ("交易单号",),
    "merchant_id": ("商户单号",),
    "note": ("备注",),
}
_DIRECTION_LABELS = {
    "income": ("收入",),
    "expense": ("支出",),
    "neutral": ("中性交易", "中性", "不计收支"),
}


def _text(value):
    return "" if value is None else str(value).strip()


def _clean_header(value):
    return re.sub(r"\s+", "", _text(value)).replace("（", "(").replace("）", ")")


def _safe_text(value):
    """Keep user-facing fields readable without payment/card details."""
    value = _ORDER_NOISE.sub("", _text(value))
    value = _NOTE_NOISE.sub("", value)
    value = _CARD_NOISE.sub("已隐藏", value)
    value = _EMAIL_NOISE.sub("已隐藏", value)
    value = re.sub(r"\s+", " ", value).strip(" ,，;；-—")
    return value


def _decimal(value):
    raw = _text(value).replace(",", "").replace("¥", "").replace("￥", "")
    raw = raw.replace("元", "").strip()
    if not raw:
        raise ValueError("empty amount")
    try:
        amount = Decimal(raw)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError("invalid amount") from exc
    if amount < 0:
        raise ValueError("negative amount")
    return amount.quantize(_ZERO, rounding=ROUND_HALF_UP)


def _money_string(value):
    return format(value.quantize(_ZERO, rounding=ROUND_HALF_UP), ".2f")


def _money_from_text(value):
    match = _MONEY.search(_text(value))
    if not match:
        raise ValueError("missing amount")
    return _decimal(match.group(0))


def _load_workbook(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read WeChat Pay XLSX statements") from exc
    try:
        return load_workbook(filename=path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("cannot read XLSX statement: %s" % exc) from exc


def _find_header(workbook):
    candidates = []
    for worksheet in workbook.worksheets:
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if row_number > 80:
                break
            values = [_clean_header(value) for value in row]
            columns = {}
            for name, aliases in _HEADERS.items():
                for alias in aliases:
                    if alias in values:
                        columns[name] = values.index(alias)
                        break
            candidates.append((len(columns), worksheet, row_number, columns))
    if not candidates:
        return None, 0
    complete_count = sum(1 for candidate in candidates if candidate[0] == len(_HEADERS))
    return max(candidates, key=lambda candidate: candidate[0]), complete_count


def detect(path):
    """Return the confidence/evidence report for a candidate XLSX file."""
    candidate = Path(path)
    evidence = []
    if candidate.suffix.lower() != ".xlsx":
        return {"format": FORMAT_ID, "confidence": 0, "evidence": ["not an .xlsx file"]}
    try:
        workbook = _load_workbook(candidate)
        found, complete_count = _find_header(workbook)
        title = "\n".join(
            _text(cell)
            for sheet in workbook.worksheets
            for row in sheet.iter_rows(max_row=12, values_only=True)
            for cell in row
            if cell is not None
        )
        workbook.close()
    except Exception as exc:
        return {"format": FORMAT_ID, "confidence": 0, "evidence": [str(exc)]}
    if not found:
        return {"format": FORMAT_ID, "confidence": 10, "evidence": ["XLSX readable; no WeChat header"]}
    if complete_count > 1:
        return {"format": FORMAT_ID, "confidence": 0, "evidence": ["multiple complete WeChat headers"]}
    count, _, row_number, columns = found
    has_wechat_marker = "微信支付" in title
    score = 10 + min(50, count * 5) + (30 if has_wechat_marker else 0)
    if count == len(_HEADERS) and has_wechat_marker:
        score = max(score, 90)
    elif count == len(_HEADERS):
        score = min(score, 70)
    evidence.extend(("header row %d" % row_number, "matched %d/%d standard columns" % (count, len(_HEADERS))))
    if has_wechat_marker:
        evidence.append("contains 微信支付 title")
    else:
        evidence.append("missing 微信支付 title marker")
    missing = sorted(set(_HEADERS) - set(columns))
    if missing:
        evidence.append("missing: " + ", ".join(missing))
    return {"format": FORMAT_ID, "confidence": min(score, 100), "evidence": evidence}


def _header_lines(worksheet, header_row):
    return [
        " ".join(_text(cell) for cell in row if _text(cell))
        for row in worksheet.iter_rows(max_row=header_row - 1, values_only=True)
    ]


def _parse_header_stats(lines):
    """Extract WeChat's preamble total and direction count/amount summaries."""
    stats = {"total": None, "income": None, "expense": None, "neutral": None}
    for line in lines:
        compact = re.sub(r"\s+", "", line)
        if stats["total"] is None:
            total = re.search(r"(?:总笔数|总计笔数|交易笔数)[:：]?(\d+)笔|共(\d+)笔记录", compact)
            if total:
                stats["total"] = int(total.group(1) or total.group(2))
        for direction, labels in _DIRECTION_LABELS.items():
            if stats[direction] is not None:
                continue
            label = "|".join(map(re.escape, labels))
            match = re.search(r"(?:%s)(?:笔数)?[:：]?(\d+)笔(.*)" % label, compact)
            if not match:
                continue
            tail = match.group(2)
            try:
                amount = _money_from_text(tail)
            except ValueError:
                amount = None
            stats[direction] = {"count": int(match.group(1)), "amount": amount}
    return stats


def _occurred_at(value):
    if isinstance(value, datetime):
        result = value
    else:
        text = _text(value)
        try:
            result = datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M"):
                try:
                    result = datetime.strptime(text, fmt)
                    break
                except ValueError:
                    result = None
            if result is None:
                raise ValueError("invalid transaction time")
    if result.tzinfo is None:
        result = result.replace(tzinfo=_TZ)
    return result.astimezone(_TZ).isoformat(timespec="seconds")


def _direction(value):
    value = _text(value)
    if "支出" in value:
        return "expense"
    if "收入" in value:
        return "income"
    if "中性" in value or "不计" in value:
        return "neutral"
    raise ValueError("unknown 收/支 value")


def _refund_from_status(status):
    if "全额退款" in status:
        return "full", None
    if "退款" not in status:
        return None, None
    amount = re.search(r"退款(?:金额)?[^\d¥￥+-]{0,12}([-+]?\s*[¥￥]?\s*[\d,]+(?:\.\d+)?)", status)
    return "partial", _money_from_text(amount.group(1)) if amount else None


def _note(transaction_type, counterparty, product, original_note, direction):
    """Return a concise merchant/product note with only non-repeated details."""
    person = _safe_text(counterparty)
    product = _safe_text(product)
    source = " ".join(part for part in (transaction_type, product) if _text(part))
    if "转账" in source:
        return "转账给" + (person or "对方")
    if "红包" in source:
        return "微信红包给" + (person or "对方")
    if person and product:
        left, right = person.casefold(), product.casefold()
        if left in right or right in left:
            return person if len(person) >= len(product) else product
        return person + " · " + product
    if person or product:
        return person or product
    fallback = _safe_text(original_note).strip(" /|·")
    if fallback:
        return fallback
    fallback = _safe_text(transaction_type).strip(" /|·")
    return fallback if fallback and fallback != "商户消费" else "微信支付交易"


def _issue(code, message, source_rows=()):
    result = {"code": code, "severity": "error", "message": message}
    if source_rows:
        result["source_rows"] = list(sorted(set(source_rows)))
    return result


def _file_sha256(path):
    digest = sha256()
    with open(path, "rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def parse(path):
    """Parse a verified WeChat export into uniform, JSON-serialisable records."""
    report = detect(path)
    if report["confidence"] < 80:
        raise ValueError("not a supported WeChat Pay XLSX statement: " + "; ".join(report["evidence"]))
    source_path = Path(path)
    file_sha256 = _file_sha256(source_path)
    workbook = _load_workbook(source_path)
    found, complete_count = _find_header(workbook)
    if complete_count > 1:
        workbook.close()
        raise ValueError("WeChat statement has multiple complete headers")
    if not found or found[0] != len(_HEADERS):
        workbook.close()
        raise ValueError("WeChat statement is missing one or more required standard columns")
    _, worksheet, header_row, columns = found
    exceptions = []
    transactions = []
    for source_row, row in enumerate(worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if not any(value is not None and _text(value) for value in row):
            continue
        def get(name):
            index = columns[name]
            return row[index] if index < len(row) else None
        if not _text(get("occurred_at")) and not _text(get("transaction_id")):
            continue  # Standard exports may include a blank footer.
        row_errors = []
        try:
            occurred_at = _occurred_at(get("occurred_at"))
        except ValueError:
            occurred_at = ""
            row_errors.append("invalid time")
        try:
            amount = _decimal(get("amount"))
        except ValueError:
            amount = _ZERO
            row_errors.append("invalid amount")
        try:
            direction = _direction(get("direction"))
        except ValueError:
            direction = "neutral"
            row_errors.append("invalid direction")
        transaction_id = _text(get("transaction_id"))
        if not transaction_id:
            row_errors.append("missing transaction id")
        status = _safe_text(get("status"))
        transaction_type = _safe_text(get("transaction_type"))
        counterparty = _safe_text(get("counterparty"))
        product = _safe_text(get("product"))
        original_note = _safe_text(get("note"))
        try:
            refund_kind, status_refund = _refund_from_status(status)
            invalid_refund_amount = False
        except ValueError:
            refund_kind, status_refund = "partial", None
            invalid_refund_amount = True
            row_errors.append("invalid refund amount")
        refund_amount = _ZERO
        exclusion_reason = None
        if direction == "income":
            exclusion_reason = "income_not_written"
        elif direction == "neutral":
            exclusion_reason = "neutral_not_written"
        elif refund_kind == "full":
            refund_amount = amount
            exclusion_reason = "fully_refunded"
        elif refund_kind == "partial":
            if not invalid_refund_amount and (status_refund is None or status_refund > amount):
                row_errors.append("invalid partial refund amount")
            elif not invalid_refund_amount:
                refund_amount = status_refund
        if amount <= _ZERO:
            row_errors.append("non-positive amount")
        if row_errors:
            exceptions.append(_issue("INVALID_TRANSACTION", "; ".join(row_errors), (source_row,)))
            exclusion_reason = exclusion_reason or "invalid_transaction"
        transaction = {
            "source_format": FORMAT_ID,
            "file_sha256": file_sha256,
            "source_row": source_row,
            "transaction_id": transaction_id,
            "transaction_key": sha256((FORMAT_ID + "\0" + file_sha256 + "\0" + (transaction_id or "row-%d" % source_row)).encode("utf-8")).hexdigest(),
            "occurred_at": occurred_at,
            "timezone": TIMEZONE,
            "currency": "CNY",
            "direction": direction,
            "original_amount": _money_string(amount),
            "refund_amount": _money_string(refund_amount),
            "net_amount": _money_string(max(_ZERO, amount - refund_amount)),
            "counterparty": counterparty,
            "product": product,
            "status": status,
            "original_note": original_note,
            "normalized_note": _note(transaction_type, counterparty, product, original_note, direction),
            "write_candidate": not row_errors and direction == "expense" and exclusion_reason is None and amount > refund_amount,
            "exclusion_reason": exclusion_reason,
            "_refund_kind": refund_kind,
            "_refund_income": direction == "income" and "退款" in (status + original_note + transaction_type + product),
        }
        transactions.append(transaction)
    header_stats = _parse_header_stats(_header_lines(worksheet, header_row))
    workbook.close()

    transaction_ids = defaultdict(list)
    for item in transactions:
        if item["transaction_id"]:
            transaction_ids[item["transaction_id"]].append(item)
    for duplicate_rows in transaction_ids.values():
        if len(duplicate_rows) < 2:
            continue
        exceptions.append(
            _issue(
                "DUPLICATE_TRANSACTION_ID",
                "the same source transaction id appears more than once",
                [item["source_row"] for item in duplicate_rows],
            )
        )
        for item in duplicate_rows:
            item["write_candidate"] = False
            item["exclusion_reason"] = "duplicate_source_transaction"

    by_direction = {name: [item for item in transactions if item["direction"] == name] for name in _DIRECTION_LABELS}
    header_invalid = header_stats["total"] is None or header_stats["total"] != len(transactions)
    if header_invalid:
        exceptions.append(_issue("HEADER_TOTAL_MISMATCH", "header total does not match parsed transactions"))
    for direction, items in by_direction.items():
        expected = header_stats[direction]
        actual_amount = sum((Decimal(item["original_amount"]) for item in items), _ZERO)
        if expected is None or expected["amount"] is None:
            exceptions.append(_issue("HEADER_%s_STATS_MISSING" % direction.upper(), "header %s count/amount is missing" % direction))
            header_invalid = True
        elif expected["count"] != len(items) or expected["amount"] != actual_amount:
            exceptions.append(_issue("HEADER_%s_MISMATCH" % direction.upper(), "header %s count/amount does not match parsed transactions" % direction))
            header_invalid = True

    expected_refunds = defaultdict(list)
    refund_income = defaultdict(list)
    for item in transactions:
        key = (_safe_text(item["counterparty"]).casefold(), item["refund_amount"])
        if item["direction"] == "expense" and item["_refund_kind"] in ("full", "partial") and item["refund_amount"] != "0.00":
            expected_refunds[key].append(item)
        elif item["_refund_income"]:
            key = (_safe_text(item["counterparty"]).casefold(), item["original_amount"])
            refund_income[key].append(item)
    unmatched_expenses = []
    unmatched_income = []
    for key in set(expected_refunds) | set(refund_income):
        expenses = expected_refunds.get(key, [])
        income = refund_income.get(key, [])
        paired = min(len(expenses), len(income))
        unmatched_expenses.extend(expenses[paired:])
        unmatched_income.extend(income[paired:])
    # WeChat can label the refund receipt with a different counterparty.  Only
    # use an amount-only fallback when one unused record exists on each side.
    expenses_by_amount = defaultdict(list)
    income_by_amount = defaultdict(list)
    for item in unmatched_expenses:
        expenses_by_amount[item["refund_amount"]].append(item)
    for item in unmatched_income:
        income_by_amount[item["original_amount"]].append(item)
    unresolved = []
    for amount in set(expenses_by_amount) | set(income_by_amount):
        expenses = expenses_by_amount.get(amount, [])
        income = income_by_amount.get(amount, [])
        if len(expenses) == len(income) == 1:
            continue
        unresolved.extend(expenses + income)
    if unresolved:
        rows = [item["source_row"] for item in unresolved]
        exceptions.append(_issue("REFUND_RECONCILIATION_MISMATCH", "refund status and refund income do not reconcile", rows))
        for item in unresolved:
            item["write_candidate"] = False
            item["exclusion_reason"] = "refund_reconciliation_mismatch" if item["direction"] == "expense" else "refund_income_mismatch"

    write_ready = not header_invalid
    for item in transactions:
        item.pop("_refund_kind", None)
        item.pop("_refund_income", None)
    return {
        "schema_version": 1,
        "source_format": FORMAT_ID,
        "file_sha256": file_sha256,
        "sheet_name": worksheet.title,
        "header_row": header_row,
        "detection": report,
        "transactions": transactions,
        "exceptions": exceptions,
        "write_ready": write_ready,
    }
